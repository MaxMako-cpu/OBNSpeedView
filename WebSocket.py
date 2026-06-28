"""
OBN SpeedView Bridge  v9
========================
One script does everything:
  - UDP listener (4DNav data)
  - WebSocket server (live data + archive + log)
  - HTTP server (serves the SpeedView HTML page to all clients)

NEW in v8:
  - LOG_DIR config field in GUI — point to current project log folder
  - Bridge reads today's "Online Log DDMMYYYY.xlsm" automatically
  - HTML sends {"type":"get_log"} via WebSocket — no file picker needed
  - UPDATE button always gets fresh data from disk (no stale file issues)

NEW in v9:
  - Two additional UDP ports for UHD333 and UHD334 node fix data
  - Node fix packets parsed: Timestamp, RL, ST, ID
  - Saved to node_fix_archive/YYYY-MM-DD.csv (both UHDs in one file)
  - Broadcast via WebSocket as {"type":"node_fix", "uhd":..., ...}

Usage:
  python obn_bridge_v9.4.py

Then everyone on the network opens:
  http://<YOUR_IP>:8080
"""

import socket
import threading
import queue
import tkinter as tk
import datetime
import time
import asyncio
import json
import logging

# Suppress noisy websockets connection errors (EOFError, InvalidMessage)
# These happen when browsers or port scanners probe the WS port — harmless
logging.getLogger("websockets.server").setLevel(logging.CRITICAL)
logging.getLogger("websockets.asyncio.server").setLevel(logging.CRITICAL)
import os
import glob
import http.server
import urllib.parse
from websockets.asyncio.server import serve as ws_serve

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── config ────────────────────────────────────────────────────────────
ARCHIVE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "speedview_archive")
os.makedirs(ARCHIVE_DIR, exist_ok=True)

NODE_FIX_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "node_fix_archive")
os.makedirs(NODE_FIX_DIR, exist_ok=True)

NODE_FIX_CSV_HEADER = "Timestamp,UHD,RL,ST,ID"

HTML_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "obn_speedview_v5.html")

CSV_HEADER = (
    "TimeStamp (Utc),"
    "IFR Easting (Metre) (32615),IFR Northing (Metre) (32615),"
    "IFR SOG (knot),IFR COG,"
    "UHD333 Easting (Metre) (32615),UHD333 Northing (Metre) (32615),"
    "UHD333 Depth (m),UHD333 Altimeter (m),UHD333 SOG (knot),UHD333 COG,"
    "TMS333_LP Range (m),TMS333_LP Vertical distance (m),TMS333_LP Declination,"
    "UHD334 Easting (Metre) (32615),UHD334 Northing (Metre) (32615),"
    "UHD334 Depth (m),UHD334 Altimeter (m),UHD334 SOG (knot),UHD334 COG,"
    "TMS334_LP Range (m),TMS334_LP Vertical distance (m),TMS334_LP Declination"
)

# ── shared state ──────────────────────────────────────────────────────
clients: set = set()
msg_queue: queue.Queue = queue.Queue()
_file_lock = threading.Lock()
_current_day = None
_current_file = None

# Node fix archive state
node_fix_queue: queue.Queue = queue.Queue()
_nf_lock = threading.Lock()
_nf_day = None
_nf_file = None

# Log directory — set via GUI
log_dir_var = None  # tk.StringVar, set in App.__init__

stats = {
    "rx_count":    0,
    "last_rx":     None,
    "fix_count":   0,
    "last_fix_rx": None,
    "last_fix":    "",
    "ws_count":    0,
    "http_port":   8080,
    "ws_port":     8765,
    "running":     False,
    "last_row":    "",
    "status":    "Stopped",
    "error":     "",
}

# ── archive helpers ───────────────────────────────────────────────────
def archive_path(day: str) -> str:
    return os.path.join(ARCHIVE_DIR, f"{day}.csv")

def save_row(csv_row: str):
    global _current_day, _current_file
    day = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    with _file_lock:
        if day != _current_day:
            if _current_file:
                try:
                    _current_file.close()
                except Exception:
                    pass
            path = archive_path(day)
            is_new = not os.path.exists(path)
            _current_file = open(path, "a", buffering=1, encoding="utf-8")
            if is_new:
                _current_file.write(CSV_HEADER + "\n")
            _current_day = day
        _current_file.write(csv_row + "\n")
        _current_file.flush()

def list_archive_days() -> list:
    files = glob.glob(os.path.join(ARCHIVE_DIR, "*.csv"))
    days = sorted([os.path.basename(f).replace(".csv", "") for f in files], reverse=True)
    return days

def read_archive_day(day: str) -> list:
    path = archive_path(day)
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if i == 0 or not line:
                continue
            rows.append(line)
    return rows

# ── Node fix helpers ──────────────────────────────────────────────────
def node_fix_path(day: str) -> str:
    return os.path.join(NODE_FIX_DIR, f"{day}.csv")

def save_node_fix(csv_row: str):
    global _nf_day, _nf_file
    day = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    with _nf_lock:
        if day != _nf_day:
            if _nf_file:
                try:
                    _nf_file.close()
                except Exception:
                    pass
            path = node_fix_path(day)
            is_new = not os.path.exists(path)
            _nf_file = open(path, "a", buffering=1, encoding="utf-8")
            if is_new:
                _nf_file.write(NODE_FIX_CSV_HEADER + "\n")
            _nf_day = day
        _nf_file.write(csv_row + "\n")
        _nf_file.flush()

def read_node_fix_day(day: str) -> list:
    """Read node fix CSV for a given day, return list of row dicts."""
    path = node_fix_path(day)
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if i == 0 or not line:
                continue
            parts = line.split(",")
            if len(parts) < 5:
                continue
            rows.append({
                "ts":  parts[0],
                "uhd": parts[1],
                "rl":  parts[2],
                "st":  parts[3],
                "id":  parts[4],
            })
    return rows

def parse_node_fix(raw: str, uhd: str):
    """
    Parse node fix UDP packet.
    Format: 2026 06 27 10:03:27.64297.00012997.00012727
    Fields separated by dots after timestamp.
    Returns dict or None.
    """
    raw = raw.strip()
    if not raw:
        return None
    try:
        # Strip leading ISO timestamp if present (e.g. "2026-06-27T07:29:25.9602338Z,")
        if "," in raw:
            raw = raw.split(",", 1)[1].strip()

        # Split by dot: ['2026 06 27 10:03:27', '64297', '00012997', '00012727']
        parts = raw.split(".")
        if len(parts) < 4:
            return None

        ts_raw = parts[0].strip()   # "2026 06 27 10:03:27"
        rl_raw = parts[1].strip()   # "64297"
        st_raw = parts[2].strip()   # "00012997"
        id_raw = parts[3].strip()   # "00012727"

        # Validate timestamp has expected shape
        if len(ts_raw) < 10:
            return None

        ts_parts = ts_raw.split(" ")
        if len(ts_parts) >= 4:
            ts = f"{ts_parts[0]}-{ts_parts[1]}-{ts_parts[2]} {ts_parts[3]}"
        else:
            ts = ts_raw

        rl = f"RL{int(rl_raw[-4:])}"
        st = f"ST{int(st_raw[-5:])}"
        id_ = f"ID{int(id_raw[-4:])}"

        return {
            "ts":  ts,
            "uhd": uhd,
            "rl":  rl,
            "st":  st,
            "id":  id_,
        }
    except Exception:
        return None

# ── Online Log reader ─────────────────────────────────────────────────
def get_log_dir() -> str:
    """Return current LOG_DIR from GUI var."""
    if log_dir_var is not None:
        return log_dir_var.get().strip()
    return ""

def find_today_log(log_dir: str) -> str:
    """Find today's log file: Online Log DDMMYYYY.xlsm"""
    if not log_dir or not os.path.isdir(log_dir):
        return ""
    now = datetime.datetime.utcnow()
    name = f"Online Log {now.strftime('%d%m%Y')}.xlsm"
    path = os.path.join(log_dir, name)
    if os.path.exists(path):
        return path
    # fallback: try without leading zero on day
    name2 = f"Online Log {now.strftime('%-d%m%Y')}.xlsm"
    path2 = os.path.join(log_dir, name2)
    if os.path.exists(path2):
        return path2
    return ""

def parse_log_xlsm(path: str, base_date_str: str) -> list:
    """
    Parse Online Log xlsm and return list of event dicts.
    base_date_str: 'YYYY-MM-DD' — the date to anchor time-of-day values.
    Mirrors the JS loadLogWorkbook logic.
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["log"] if "log" in wb.sheetnames else wb.active

    # Find header row (look for "REMARKS" in first 40 rows)
    header_row = 16  # default
    for i, row in enumerate(ws.iter_rows(max_row=40, values_only=True)):
        for cell in (row or []):
            if isinstance(cell, str) and cell.strip().upper() == "REMARKS":
                header_row = i
                break
        else:
            continue
        break

    # Parse base date
    try:
        base_dt = datetime.datetime.strptime(base_date_str, "%Y-%m-%d")
    except Exception:
        base_dt = datetime.datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)

    base_epoch = base_dt.timestamp()

    events = []
    prev_ms = None
    day_offset = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row + 1:
            continue
        if not row:
            continue

        # Column layout (0-indexed): col5=UTC time, col4=code, col7=remark
        utc_start = row[5] if len(row) > 5 else None
        code      = row[4] if len(row) > 4 else None
        remark    = row[7] if len(row) > 7 else None

        if utc_start is None:
            continue

        # Convert time value to ms-of-day
        ms = None
        if isinstance(utc_start, datetime.time):
            ms = (utc_start.hour * 3600 + utc_start.minute * 60 + utc_start.second) * 1000
        elif isinstance(utc_start, datetime.datetime):
            ms = (utc_start.hour * 3600 + utc_start.minute * 60 + utc_start.second) * 1000
        elif isinstance(utc_start, (int, float)):
            ms = round((utc_start - int(utc_start)) * 86400000)
        elif isinstance(utc_start, str):
            parts = utc_start.strip().split(":")
            if len(parts) >= 2:
                try:
                    h, m = int(parts[0]), int(parts[1])
                    s = int(float(parts[2])) if len(parts) > 2 else 0
                    ms = (h * 3600 + m * 60 + s) * 1000
                except Exception:
                    pass

        if ms is None:
            continue

        # Midnight rollover detection
        if prev_ms is not None and ms < prev_ms - 60000:
            day_offset += 1
        prev_ms = ms

        text = str(remark).strip() if remark is not None else ""
        if not text:
            continue

        epoch = base_epoch + day_offset * 86400 + ms / 1000
        events.append({
            "epoch":  round(epoch, 3),
            "code":   str(code).strip() if code is not None else "",
            "remark": text,
        })

    wb.close()
    events.sort(key=lambda e: e["epoch"])
    return events

# ── UDP parser ────────────────────────────────────────────────────────
def parse_udp_to_csv(raw: str):
    line = raw.strip()
    if not line:
        return None
    if "*" in line:
        line = line[:line.rfind("*")]
    line = line.strip()
    parts = line.split(",")
    if len(parts) < 22:
        return None
    try:
        float(parts[0])
    except ValueError:
        return None
    try:
        dt  = datetime.datetime.utcnow()
        ts  = dt.strftime("%Y-%m-%d %H:%M:%S.") + f"{dt.microsecond // 1000:03d}"
        f   = parts
        def pos(v):
            try:
                return str(abs(float(v)))
            except Exception:
                return "N/A"
        # 4DNav 22-field format (Speed Web Log):
        # 0  IFR Easting          1  IFR Northing
        # 2  IFR SOG              3  IFR COG
        # 4  UHD333 Easting       5  UHD333 Northing
        # 6  UHD333 Depth(P2D)    7  UHD333 COG
        # 8  TMS333 LP Range      9  UHD333 SOG
        # 10 TMS333 LP VDist      11 TMS333 LP Decl
        # 12 UHD334 Easting       13 UHD334 Northing
        # 14 UHD334 SOG           15 UHD334 COG
        # 16 TMS334 LP Range      17 TMS334 LP VDist
        # 18 TMS334 LP Decl       19 UHD334 Alt(LNAV)
        # 20 UHD333 Alt(LNAV)     21 UHD334 Depth(P2D)
        return ",".join([
            ts,
            f[0],  f[1],          # IFR Easting, Northing
            f[2],  f[3],          # IFR SOG, COG
            f[4],  f[5],          # UHD333 Easting, Northing
            pos(f[6]), f[20],     # UHD333 Depth, Altimeter(LNAV)
            f[9],  f[7],          # UHD333 SOG, COG
            f[8],  f[10], f[11],  # TMS333 LP Range, VDist, Decl
            f[12], f[13],         # UHD334 Easting, Northing
            pos(f[21]), f[19],    # UHD334 Depth, Altimeter(LNAV)
            f[14], f[15],         # UHD334 SOG, COG
            f[16], f[17], f[18],  # TMS334 LP Range, VDist, Decl
        ])
    except (ValueError, IndexError):
        return None

# ── UDP Listener ──────────────────────────────────────────────────────
class UdpListener:
    def __init__(self, port, on_data, on_status):
        self.port      = port
        self.on_data   = on_data
        self.on_status = on_status
        self._running  = False
        self._sock     = None

    def start(self):
        self._running = True
        threading.Thread(target=self._run, daemon=True).start()

    def stop(self):
        self._running = False
        if self._sock:
            try:
                self._sock.close()
            except Exception:
                pass

    def _run(self):
        try:
            self._sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            self._sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            self._sock.bind(("0.0.0.0", self.port))
            self._sock.settimeout(1.0)
            self.on_status(f"Listening on UDP :{self.port}", "ok")
        except Exception as e:
            self.on_status(f"Bind error: {e}", "error")
            self._running = False
            return
        while self._running:
            try:
                raw, _ = self._sock.recvfrom(4096)
                self.on_data(raw.decode("ascii", errors="replace").strip())
            except socket.timeout:
                continue
            except Exception as e:
                if self._running:
                    self.on_status(f"Socket error: {e}", "error")
                break

# ── HTTP server ───────────────────────────────────────────────────────
class SpeedViewHandler(http.server.BaseHTTPRequestHandler):
    def do_GET(self):
        path = urllib.parse.urlparse(self.path).path
        if path in ("/", "/index.html", "/obn_speedview_v5.html"):
            if not os.path.exists(HTML_FILE):
                self.send_error(404, "obn_speedview_v5.html not found next to bridge script")
                return
            with open(HTML_FILE, "rb") as f:
                data = f.read()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        else:
            self.send_error(404)

    def log_message(self, format, *args):
        pass

def start_http_server(port: int):
    server = http.server.HTTPServer(("0.0.0.0", port), SpeedViewHandler)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    return server

# ── WebSocket ─────────────────────────────────────────────────────────
async def ws_handler(websocket):
    clients.add(websocket)
    stats["ws_count"] = len(clients)
    try:
        await websocket.send(json.dumps({"type": "header", "row": CSV_HEADER}))
        async for message in websocket:
            try:
                req   = json.loads(message)
                rtype = req.get("type")

                if rtype == "get_days":
                    await websocket.send(json.dumps({
                        "type": "days",
                        "list": list_archive_days()
                    }))

                elif rtype == "get_archive":
                    day  = req.get("day", "")
                    rows = read_archive_day(day)
                    await websocket.send(json.dumps({
                        "type": "archive",
                        "day":  day,
                        "rows": rows
                    }))

                elif rtype == "get_node_fixes":
                    day  = req.get("day", "")
                    rows = read_node_fix_day(day)
                    await websocket.send(json.dumps({
                        "type": "node_fixes_archive",
                        "day":  day,
                        "rows": rows
                    }))

                elif rtype == "get_log":
                    # Read today's Online Log from LOG_DIR
                    base_date = req.get("base_date", "")  # YYYY-MM-DD sent by client
                    ldir = get_log_dir()
                    if not ldir:
                        await websocket.send(json.dumps({
                            "type":  "log_error",
                            "error": "Log directory not set. Please set LOG DIR in bridge GUI."
                        }))
                        continue
                    log_path = find_today_log(ldir)
                    if not log_path:
                        now = datetime.datetime.utcnow()
                        expected = f"Online Log {now.strftime('%d%m%Y')}.xlsm"
                        await websocket.send(json.dumps({
                            "type":  "log_error",
                            "error": f"Log file not found: {expected}\nIn folder: {ldir}"
                        }))
                        continue
                    try:
                        events = parse_log_xlsm(log_path, base_date)
                        await websocket.send(json.dumps({
                            "type":   "log_data",
                            "events": events,
                            "file":   os.path.basename(log_path),
                            "count":  len(events)
                        }))
                    except Exception as e:
                        await websocket.send(json.dumps({
                            "type":  "log_error",
                            "error": str(e)
                        }))

            except Exception:
                pass
    except Exception:
        pass
    finally:
        clients.discard(websocket)
        stats["ws_count"] = len(clients)

async def broadcaster():
    while stats["running"]:
        dead = set()
        # ── main speed data ──
        while not msg_queue.empty():
            try:
                row = msg_queue.get_nowait()
                msg = json.dumps({"type": "live", "row": row})
                for ws in list(clients):
                    try:
                        await ws.send(msg)
                    except Exception:
                        dead.add(ws)
            except queue.Empty:
                break
        # ── node fix data ──
        while not node_fix_queue.empty():
            try:
                fix = node_fix_queue.get_nowait()
                msg = json.dumps({"type": "node_fix",
                                  "uhd": fix["uhd"],
                                  "ts":  fix["ts"],
                                  "rl":  fix["rl"],
                                  "st":  fix["st"],
                                  "id":  fix["id"]})
                for ws in list(clients):
                    try:
                        await ws.send(msg)
                    except Exception:
                        dead.add(ws)
            except queue.Empty:
                break
        clients.difference_update(dead)
        stats["ws_count"] = len(clients)
        await asyncio.sleep(0.05)

async def ws_main(ws_port: int):
    async with ws_serve(ws_handler, "0.0.0.0", ws_port):
        await broadcaster()

# ── UDP callbacks ─────────────────────────────────────────────────────
def on_udp_data(text: str):
    csv_row = parse_udp_to_csv(text)
    if csv_row:
        stats["rx_count"] += 1
        stats["last_rx"]   = time.time()
        stats["last_row"]  = csv_row[:90] + "…"
        save_row(csv_row)
        msg_queue.put(csv_row)

def on_udp_status(msg: str, level: str):
    if level == "error":
        stats["status"] = msg
        stats["error"]  = msg
    elif not stats["running"]:
        stats["status"] = msg

def make_node_fix_callbacks(uhd: str):
    """Factory — returns (on_data, on_status) for given UHD label."""
    def on_data(text: str):
        fix = parse_node_fix(text, uhd)
        if fix:
            csv_row = f"{fix['ts']},{fix['uhd']},{fix['rl']},{fix['st']},{fix['id']}"
            save_node_fix(csv_row)
            node_fix_queue.put(fix)
            stats["fix_count"]   += 1
            stats["last_fix_rx"]  = time.time()
            stats["last_fix"]     = f"{uhd}: {fix['rl']} {fix['st']} {fix['id']} @ {fix['ts']}"
        else:
            # Log raw unparsed packet so we can debug format issues
            stats["last_fix"] = f"{uhd} RAW (parse fail): {text[:80]}"
            stats["last_fix_rx"] = time.time()
    def on_status(msg: str, level: str):
        if level == "error":
            stats["error"] = f"{uhd} fix: {msg}"
    return on_data, on_status

# ── bridge start/stop ─────────────────────────────────────────────────
_listener    = None
_ws_loop     = None
_ws_thread   = None
_http_server = None
_listener333 = None
_listener334 = None

def start_bridge(udp_port: int, ws_port: int, http_port: int,
                 fix333_port: int = 0, fix334_port: int = 0):
    global _listener, _ws_loop, _ws_thread, _http_server, _listener333, _listener334
    stats["running"]   = True
    stats["rx_count"]  = 0
    stats["error"]     = ""
    stats["last_row"]  = ""
    stats["ws_port"]   = ws_port
    stats["http_port"] = http_port

    _listener = UdpListener(udp_port, on_udp_data, on_udp_status)
    _listener.start()

    if fix333_port > 0:
        on_data333, on_status333 = make_node_fix_callbacks("UHD333")
        _listener333 = UdpListener(fix333_port, on_data333, on_status333)
        _listener333.start()

    if fix334_port > 0:
        on_data334, on_status334 = make_node_fix_callbacks("UHD334")
        _listener334 = UdpListener(fix334_port, on_data334, on_status334)
        _listener334.start()

    try:
        _http_server = start_http_server(http_port)
    except Exception as e:
        stats["error"] = f"HTTP error: {e}"

    def run_ws():
        global _ws_loop
        _ws_loop = asyncio.new_event_loop()
        asyncio.set_event_loop(_ws_loop)
        try:
            _ws_loop.run_until_complete(ws_main(ws_port))
        except Exception as e:
            stats["error"] = f"WS error: {e}"
        finally:
            _ws_loop.close()

    _ws_thread = threading.Thread(target=run_ws, daemon=True)
    _ws_thread.start()

def stop_bridge():
    stats["running"] = False
    if _listener:
        _listener.stop()
    if _listener333:
        _listener333.stop()
    if _listener334:
        _listener334.stop()
    if _http_server:
        threading.Thread(target=_http_server.shutdown, daemon=True).start()
    if _ws_loop and not _ws_loop.is_closed():
        _ws_loop.call_soon_threadsafe(_ws_loop.stop)

# ── GUI ───────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("OBN SpeedView Bridge v9.4")
        self.resizable(False, False)
        self.configure(bg="#0c1622")

        FG    = "#d8e6f2"
        BG    = "#0c1622"
        PANEL = "#0e1a28"
        CYAN  = "#7fd4ff"
        GREEN = "#3ee07a"
        AMBER = "#ffb454"
        MONO  = ("Consolas", 10)

        # ── title ──
        tk.Label(self, text="OBN SPEEDVIEW  ·  BRIDGE  v9.4",
                 bg=BG, fg=CYAN, font=("Consolas", 12, "bold"),
                 pady=10).pack(fill="x")

        # ── port config frame ──
        frame_cfg = tk.Frame(self, bg=PANEL, pady=6)
        frame_cfg.pack(fill="x", padx=10, pady=(0, 4))

        for col, (label, attr, default) in enumerate([
            ("UDP",        "udp_var",    "7004"),
            ("WebSocket",  "ws_var",     "8765"),
            ("HTTP",       "http_var",   "8080"),
            ("Fix333",     "fix333_var", "7005"),
            ("Fix334",     "fix334_var", "7006"),
        ]):
            tk.Label(frame_cfg, text=label, bg=PANEL, fg=FG,
                     font=MONO).grid(row=0, column=col*2, padx=(10, 2), pady=4, sticky="w")
            var = tk.StringVar(value=default)
            setattr(self, attr, var)
            tk.Entry(frame_cfg, textvariable=var, width=6,
                     bg="#101724", fg=CYAN, insertbackground=CYAN,
                     font=MONO, relief="flat", bd=4
                     ).grid(row=0, column=col*2+1, padx=(2, 6), pady=4)

        # ── LOG DIR frame ──
        frame_log = tk.Frame(self, bg=PANEL)
        frame_log.pack(fill="x", padx=10, pady=(0, 4))

        tk.Label(frame_log, text="LOG DIR:", bg=PANEL, fg=AMBER,
                 font=MONO).grid(row=0, column=0, padx=(10, 4), pady=6, sticky="w")

        global log_dir_var
        log_dir_var = tk.StringVar(value="")
        self.log_dir_entry = tk.Entry(
            frame_log, textvariable=log_dir_var, width=48,
            bg="#101724", fg=AMBER, insertbackground=AMBER,
            font=MONO, relief="flat", bd=4)
        self.log_dir_entry.grid(row=0, column=1, padx=4, pady=6, sticky="ew")

        tk.Button(frame_log, text="Browse…", command=self._browse_log_dir,
                  bg="#1a2030", fg=CYAN, font=MONO, relief="flat",
                  padx=8, pady=3, cursor="hand2"
                  ).grid(row=0, column=2, padx=(4, 10), pady=6)

        self.log_status_var = tk.StringVar(value="No log directory set")
        tk.Label(frame_log, textvariable=self.log_status_var,
                 bg=PANEL, fg="#6f8aa3", font=("Consolas", 9), anchor="w"
                 ).grid(row=1, column=0, columnspan=3, padx=10, pady=(0, 4), sticky="w")

        frame_log.columnconfigure(1, weight=1)

        # ── START / STOP buttons ──
        frame_btn = tk.Frame(self, bg=BG)
        frame_btn.pack(fill="x", padx=10, pady=6)

        self.btn_start = tk.Button(
            frame_btn, text="▶  START", command=self.on_start,
            bg="#0a2b30", fg=CYAN, font=("Consolas", 10, "bold"),
            relief="flat", padx=16, pady=6, cursor="hand2")
        self.btn_start.pack(side="left", expand=True, fill="x", padx=(0, 4))

        self.btn_stop = tk.Button(
            frame_btn, text="■  STOP", command=self.on_stop,
            bg="#1a0e0e", fg="#ff9a9a", font=("Consolas", 10, "bold"),
            relief="flat", padx=16, pady=6, cursor="hand2", state="disabled")
        self.btn_stop.pack(side="left", expand=True, fill="x", padx=(4, 0))

        # ── status panel ──
        frame_st = tk.Frame(self, bg=PANEL)
        frame_st.pack(fill="x", padx=10, pady=4)

        # Status dot + text
        row0 = tk.Frame(frame_st, bg=PANEL)
        row0.pack(fill="x", padx=10, pady=(8, 2))

        self.dot_canvas = tk.Canvas(row0, width=16, height=16,
                                    bg=PANEL, highlightthickness=0)
        self.dot_canvas.pack(side="left", padx=(0, 8))
        self.dot = self.dot_canvas.create_oval(2, 2, 14, 14, fill="#3d5168", outline="")

        self.status_var = tk.StringVar(value="Stopped")
        tk.Label(row0, textvariable=self.status_var,
                 bg=PANEL, fg=FG, font=MONO, anchor="w").pack(side="left")

        # Stats grid
        frame_stats = tk.Frame(frame_st, bg=PANEL)
        frame_stats.pack(fill="x", padx=10, pady=2)

        for i, (label, color) in enumerate([
            ("UDP packets rx:", GREEN),
            ("Node fixes rx:",  "#c792ea"),
            ("WS clients:",     CYAN),
            ("Archive days:",   AMBER),
        ]):
            tk.Label(frame_stats, text=label, bg=PANEL, fg="#6f8aa3",
                     font=MONO).grid(row=i, column=0, sticky="w", pady=1)
            var = tk.StringVar(value="0")
            setattr(self, f"_var_{i+1}", var)
            tk.Label(frame_stats, textvariable=var, bg=PANEL, fg=color,
                     font=("Consolas", 10, "bold"), anchor="w",
                     ).grid(row=i, column=1, sticky="w", padx=8)

        # Last main packet
        tk.Label(frame_st, text="Last packet:", bg=PANEL, fg="#6f8aa3",
                 font=MONO, anchor="w").pack(fill="x", padx=10, pady=(6, 0))
        self.last_var = tk.StringVar(value="—")
        tk.Label(frame_st, textvariable=self.last_var, bg=PANEL, fg=AMBER,
                 font=("Consolas", 9), anchor="w", wraplength=560, justify="left"
                 ).pack(fill="x", padx=10, pady=(0, 2))

        # Last node fix
        tk.Label(frame_st, text="Last node fix:", bg=PANEL, fg="#6f8aa3",
                 font=MONO, anchor="w").pack(fill="x", padx=10, pady=(4, 0))
        self.last_fix_var = tk.StringVar(value="—")
        tk.Label(frame_st, textvariable=self.last_fix_var, bg=PANEL, fg="#c792ea",
                 font=("Consolas", 9), anchor="w", wraplength=560, justify="left"
                 ).pack(fill="x", padx=10, pady=(0, 4))

        self.err_var = tk.StringVar(
            value="" if OPENPYXL_OK
            else "⚠ openpyxl not installed — log reading disabled.")
        tk.Label(frame_st, textvariable=self.err_var, bg=PANEL, fg="#ff4d6a",
                 font=("Consolas", 9), anchor="w", wraplength=560
                 ).pack(fill="x", padx=10, pady=(0, 8))

        # ── URL display ──
        self.url_var = tk.StringVar(value="")
        tk.Label(self, textvariable=self.url_var, bg=BG, fg=GREEN,
                 font=("Consolas", 10, "bold"), pady=6).pack()

        tk.Label(self, text=f"Archive: {ARCHIVE_DIR}",
                 bg=BG, fg="#3d5168", font=("Consolas", 7), pady=1).pack()
        tk.Label(self, text=f"Node fixes: {NODE_FIX_DIR}",
                 bg=BG, fg="#3d5168", font=("Consolas", 7), pady=1).pack()

        self._dot_state = False

        # Watch log dir for changes
        log_dir_var.trace_add("write", self._on_log_dir_change)
        self._update_ui()

    def _browse_log_dir(self):
        from tkinter import filedialog
        d = filedialog.askdirectory(title="Select Online Log folder")
        if d:
            log_dir_var.set(d)

    def _on_log_dir_change(self, *_):
        d = log_dir_var.get().strip()
        if not d:
            self.log_status_var.set("No log directory set")
            return
        if not os.path.isdir(d):
            self.log_status_var.set("⚠ Directory not found")
            return
        now = datetime.datetime.utcnow()
        expected = f"Online Log {now.strftime('%d%m%Y')}.xlsm"
        full = os.path.join(d, expected)
        if os.path.exists(full):
            self.log_status_var.set(f"✓ Found: {expected}")
        else:
            self.log_status_var.set(f"Today's log not found yet: {expected}")

    def on_start(self):
        try:
            udp  = int(self.udp_var.get())
            ws   = int(self.ws_var.get())
            http = int(self.http_var.get())
            assert all(1 < p < 65535 for p in [udp, ws, http])
        except Exception:
            self.status_var.set("Invalid port numbers!")
            return
        try:
            fix333 = int(self.fix333_var.get()) if self.fix333_var.get().strip() else 0
        except Exception:
            fix333 = 0
        try:
            fix334 = int(self.fix334_var.get()) if self.fix334_var.get().strip() else 0
        except Exception:
            fix334 = 0
        if not os.path.exists(HTML_FILE):
            self.err_var.set(f"⚠ HTML not found: {HTML_FILE}")
        start_bridge(udp, ws, http, fix333, fix334)
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
        except Exception:
            ip = "YOUR_IP"
        self.url_var.set(f"▶  http://{ip}:{http}  (share this with your team)")
        ports_info = f"UDP :{udp}"
        if fix333 > 0:
            ports_info += f"  |  Fix333 :{fix333}"
        if fix334 > 0:
            ports_info += f"  |  Fix334 :{fix334}"
        self.status_var.set(f"Listening — {ports_info}")

    def on_stop(self):
        stop_bridge()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.status_var.set("Stopped")
        self.url_var.set("")
        self.dot_canvas.itemconfig(self.dot, fill="#3d5168")

    def _update_ui(self):
        self._var_1.set(f"{stats['rx_count']:,}")
        self._var_2.set(f"{stats['fix_count']:,}")
        self._var_3.set(str(stats["ws_count"]))
        self._var_4.set(str(len(list_archive_days())))
        if stats["last_row"]:
            self.last_var.set(stats["last_row"])
        if stats["last_fix"]:
            self.last_fix_var.set(stats["last_fix"])
        if stats["error"]:
            self.err_var.set(stats["error"])
            stats["error"] = ""

        now = time.time()
        if stats["last_fix_rx"] and (now - stats["last_fix_rx"]) < 2:
            self._dot_state = not self._dot_state
            self.dot_canvas.itemconfig(self.dot,
                fill="#c792ea" if self._dot_state else "#5a3070")
        elif stats["last_rx"] and (now - stats["last_rx"]) < 2:
            self._dot_state = not self._dot_state
            self.dot_canvas.itemconfig(self.dot,
                fill="#3ee07a" if self._dot_state else "#1a5c30")
        elif stats["running"]:
            self.dot_canvas.itemconfig(self.dot, fill="#ffb454")
        else:
            self.dot_canvas.itemconfig(self.dot, fill="#3d5168")

        self.after(500, self._update_ui)

    def on_closing(self):
        stop_bridge()
        self.destroy()

if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()